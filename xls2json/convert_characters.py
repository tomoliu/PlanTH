import json, os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(__file__))
INPUT = os.path.join(BASE, "data", "tables", "角色表.xlsx")
OUTPUT = os.path.join(BASE, "data", "json", "角色表.json")

wb = load_workbook(INPUT, data_only=True)
ws = wb.active

characters: list[dict] = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if len(row) < 3 or row[0] is None:
        continue
    characters.append({
        "id": int(row[0]),
        "name": str(row[1]),
        "lobby_background": str(row[2]) if row[2] else "",
        "initial_hp": int(row[3]) if len(row) > 3 and row[3] is not None else 5,
        "initial_gold": int(row[4]) if len(row) > 4 and row[4] is not None else 50,
    })

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"characters": characters}, f, ensure_ascii=False, indent=2)
print(f"Wrote {len(characters)} characters to {OUTPUT}")
