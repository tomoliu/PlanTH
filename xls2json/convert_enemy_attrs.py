import json, os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(__file__))
INPUT = os.path.join(BASE, "data", "tables", "敌人属性表.xlsx")
OUTPUT = os.path.join(BASE, "data", "json", "敌人属性表.json")

wb = load_workbook(INPUT, data_only=True)
ws = wb.active

enemies: list[dict] = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    enemies.append({
        "id": int(row[0]),
        "name": str(row[1]) if row[1] else "",
        "speed": float(row[2]) if len(row) > 2 and row[2] is not None else 240.0,
        "hp": int(row[3]) if len(row) > 3 and row[3] is not None else 3,
        "fire_interval": float(row[4]) if len(row) > 4 and row[4] is not None else 2.0,
        "contact_damage": int(row[5]) if len(row) > 5 and row[5] is not None else 1,
        "bullet_speed": float(row[6]) if len(row) > 6 and row[6] is not None else 250.0,
    })

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"enemies": enemies}, f, ensure_ascii=False, indent=2)
print(f"Wrote {len(enemies)} enemies to {OUTPUT}")
