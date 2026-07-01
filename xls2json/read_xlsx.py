from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\Administrator\Documents\PlanTH\data\tables\角色表.xlsx", data_only=True)
ws = wb.active
print("Sheet:", ws.title)
print("Rows:", ws.max_row, "Cols:", ws.max_column)
for row in ws.iter_rows(values_only=True):
    print(row)
