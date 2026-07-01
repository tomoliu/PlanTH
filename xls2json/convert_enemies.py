import json, os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(__file__))
INPUT = os.path.join(BASE, "data", "tables", "敌人配置表.xlsx")
OUTPUT = os.path.join(BASE, "data", "json", "敌人配置表.json")

wb = load_workbook(INPUT, data_only=True)
ws = wb.active

levels: list[dict] = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    levels.append({
        "difficulty": int(row[0]),
        "enemy_count": int(row[1]),
        "condition": str(row[2]) if len(row) > 2 and row[2] else "",
    })

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump({"levels": levels}, f, ensure_ascii=False, indent=2)
print(f"Wrote {len(levels)} levels to {OUTPUT}")
