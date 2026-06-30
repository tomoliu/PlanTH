import json
from openpyxl import load_workbook

INPUT_PATH = r"C:\Users\Administrator\Documents\PlanTH\data\tables\角色表.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Documents\PlanTH\data\tables\角色表.json"

wb = load_workbook(INPUT_PATH, data_only=True)
ws = wb.active

characters: list[dict] = []
# Skip header row (row 1)
for row in ws.iter_rows(min_row=2, values_only=True):
    char_id, name, lobby_background = row
    if char_id is None:
        continue
    characters.append({
        "id": int(char_id),
        "name": str(name),
        "lobby_background": str(lobby_background)
    })

with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump({"characters": characters}, f, ensure_ascii=False, indent=2)

print(f"Wrote {len(characters)} characters to {OUTPUT_PATH}")
